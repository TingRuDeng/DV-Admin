# -*- coding: utf-8 -*-
"""用户 Excel 导入、模板下载和 CSV 导出服务。"""

from __future__ import annotations

import base64
import csv
import io
from dataclasses import dataclass
from typing import Any, BinaryIO

from django.conf import settings
from django.db import transaction
from rest_framework.exceptions import ValidationError

from drf_admin.apps.system.models import Departments, Roles, Users
from drf_admin.apps.system.services.data_scope import (
    apply_user_data_scope,
    get_visible_department_ids,
)
from drf_admin.apps.system.services.field_permission import (
    USER_FIELD_PLAIN_PERMISSION,
    can_view_plain_fields,
    can_write_sensitive_user_fields,
    mask_email,
    mask_mobile,
)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_CONTENT_TYPE = "text/csv;charset=utf-8"


@dataclass(frozen=True)
class ImportColumns:
    username: int
    name: int | None
    email: int | None
    mobile: int | None
    gender: int | None
    dept: int | None
    role: int | None


@dataclass
class ImportContext:
    all_depts: dict[int, Departments]
    all_roles: dict[int, Roles]
    existing_usernames: set[str]
    existing_mobiles: set[str]


@dataclass(frozen=True)
class ImportRow:
    username: str
    name: str
    email: str | None
    mobile: str | None
    gender: int
    dept_id: int | None
    role_ids: tuple[int, ...]


def build_import_template() -> dict[str, str]:
    """生成与 FastAPI 相同列定义的 xlsx 模板。"""
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户导入模板"
    worksheet.append(
        ["用户名*", "姓名", "邮箱", "手机号", "性别", "部门ID", "角色ID(多个用逗号分隔)"]
    )
    worksheet.append(
        ["zhangsan", "张三", "zhangsan@example.com", "13800138000", "1", "1", "1,2"]
    )
    for column, width in {
        "A": 15,
        "B": 15,
        "C": 25,
        "D": 15,
        "E": 10,
        "F": 10,
        "G": 25,
    }.items():
        worksheet.column_dimensions[column].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return _encoded_file("用户导入模板.xlsx", buffer.getvalue(), XLSX_CONTENT_TYPE)


def export_users(current_user: Users) -> dict[str, str]:
    """导出当前用户数据范围内的 CSV，并按字段权限脱敏。"""
    queryset = apply_user_data_scope(Users.objects.all(), current_user).order_by("id")
    can_view_plain = can_view_plain_fields(current_user, USER_FIELD_PLAIN_PERMISSION)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["用户名", "姓名", "邮箱", "手机号", "状态", "创建时间"])
    for user in queryset:
        writer.writerow(
            [
                user.username,
                user.name or "",
                (user.email if can_view_plain else mask_email(user.email)) or "",
                (user.mobile if can_view_plain else mask_mobile(user.mobile)) or "",
                "启用" if user.is_active else "禁用",
                user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "",
            ]
        )
    content = ("\ufeff" + buffer.getvalue()).encode("utf-8")
    return _encoded_file("用户导出.csv", content, CSV_CONTENT_TYPE)


@transaction.atomic
def import_users(
    file: BinaryIO,
    *,
    dept_id: int | None,
    current_user: Users,
) -> dict[str, Any]:
    """逐行导入用户；无效行返回明细，意外异常则回滚本次写入。"""
    worksheet, workbook = _load_worksheet(file)
    try:
        columns = _parse_columns(worksheet)
        context = _build_context()
        visible_dept_ids = get_visible_department_ids(current_user)
        can_write_sensitive = can_write_sensitive_user_fields(current_user)
        valid_count = 0
        invalid_count = 0
        messages: list[str] = []

        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            parsed = _parse_row(
                row_idx,
                row,
                columns,
                dept_id,
                context,
                messages,
                visible_dept_ids,
                can_write_sensitive,
            )
            if parsed is None:
                invalid_count += 1
                continue
            user = Users.objects.create_user(
                username=parsed.username,
                password=settings.DEFAULT_PWD,
                name=parsed.name,
                email=parsed.email,
                mobile=parsed.mobile,
                gender=parsed.gender,
                is_active=1,
                dept_id=parsed.dept_id,
            )
            if parsed.role_ids:
                user.roles.set([context.all_roles[role_id] for role_id in parsed.role_ids])
            valid_count += 1
    finally:
        workbook.close()

    return {
        "validCount": valid_count,
        "invalidCount": invalid_count,
        "messageList": messages,
    }


def _encoded_file(filename: str, content: bytes, content_type: str) -> dict[str, str]:
    return {
        "filename": filename,
        "content": base64.b64encode(content).decode("ascii"),
        "contentType": content_type,
    }


def _load_worksheet(file: BinaryIO):
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Excel 文件解析失败: {exc}") from exc
    return workbook.active, workbook


def _parse_columns(worksheet) -> ImportColumns:
    headers = [cell.value for cell in worksheet[1]]
    if not headers or headers[0] is None:
        raise ValidationError("Excel 文件格式错误，缺少表头")
    if "用户名*" not in headers:
        raise ValidationError("Excel 文件缺少必要字段: 用户名*")
    return ImportColumns(
        username=headers.index("用户名*"),
        name=headers.index("姓名") if "姓名" in headers else None,
        email=headers.index("邮箱") if "邮箱" in headers else None,
        mobile=headers.index("手机号") if "手机号" in headers else None,
        gender=headers.index("性别") if "性别" in headers else None,
        dept=headers.index("部门ID") if "部门ID" in headers else None,
        role=(
            headers.index("角色ID(多个用逗号分隔)")
            if "角色ID(多个用逗号分隔)" in headers
            else None
        ),
    )


def _build_context() -> ImportContext:
    return ImportContext(
        all_depts={dept.id: dept for dept in Departments.objects.all()},
        all_roles={role.id: role for role in Roles.objects.filter(status=1)},
        existing_usernames=set(Users.objects.values_list("username", flat=True)),
        existing_mobiles={
            str(mobile)
            for mobile in Users.objects.exclude(mobile__isnull=True).values_list(
                "mobile", flat=True
            )
            if mobile
        },
    )


def _parse_row(
    row_idx: int,
    row: tuple[Any, ...],
    columns: ImportColumns,
    default_dept_id: int | None,
    context: ImportContext,
    messages: list[str],
    visible_dept_ids: set[int] | None,
    can_write_sensitive: bool,
) -> ImportRow | None:
    username = _optional_text(row, columns.username)
    if not username:
        messages.append(f"第{row_idx}行: 用户名不能为空")
        return None
    if username in context.existing_usernames:
        messages.append(f"第{row_idx}行: 用户名 '{username}' 已存在")
        return None

    email = _optional_text(row, columns.email)
    mobile = _optional_text(row, columns.mobile)
    if mobile and mobile in context.existing_mobiles:
        messages.append(f"第{row_idx}行: 手机号 '{mobile}' 已存在")
        return None
    if not can_write_sensitive and any((email, mobile)):
        messages.append(f"第{row_idx}行: 缺少字段写入权限，不能写入手机号或邮箱")
        return None

    role_ids = _parse_role_ids(row_idx, row, columns.role, context, messages)
    if role_ids is None:
        return None
    resolved_dept_id = _parse_dept_id(
        row_idx,
        row,
        columns.dept,
        default_dept_id,
        context,
        messages,
    )
    if visible_dept_ids is not None and resolved_dept_id not in visible_dept_ids:
        messages.append(f"第{row_idx}行: 目标部门超出当前用户数据范围")
        return None

    context.existing_usernames.add(username)
    if mobile:
        context.existing_mobiles.add(mobile)
    return ImportRow(
        username=username,
        name=_optional_text(row, columns.name) or username,
        email=email,
        mobile=mobile,
        gender=_parse_gender(row, columns.gender),
        dept_id=resolved_dept_id,
        role_ids=role_ids,
    )


def _parse_dept_id(
    row_idx: int,
    row: tuple[Any, ...],
    column: int | None,
    default_dept_id: int | None,
    context: ImportContext,
    messages: list[str],
) -> int | None:
    value = row[column] if column is not None else None
    if value is None:
        return default_dept_id
    try:
        imported_dept_id = int(value)
    except (TypeError, ValueError):
        messages.append(f"第{row_idx}行: 部门ID格式错误，使用默认部门")
        return default_dept_id
    if imported_dept_id not in context.all_depts:
        messages.append(f"第{row_idx}行: 部门ID '{imported_dept_id}' 不存在，使用默认部门")
        return default_dept_id
    return imported_dept_id


def _parse_role_ids(
    row_idx: int,
    row: tuple[Any, ...],
    column: int | None,
    context: ImportContext,
    messages: list[str],
) -> tuple[int, ...] | None:
    value = row[column] if column is not None else None
    if value is None:
        return ()
    try:
        role_ids = tuple(
            dict.fromkeys(int(item.strip()) for item in str(value).split(",") if item.strip())
        )
    except (TypeError, ValueError):
        messages.append(f"第{row_idx}行: 角色ID格式错误")
        return None
    invalid_role_ids = [role_id for role_id in role_ids if role_id not in context.all_roles]
    if invalid_role_ids:
        messages.append(f"第{row_idx}行: 角色ID {invalid_role_ids} 不存在或已禁用")
        return None
    return role_ids


def _parse_gender(row: tuple[Any, ...], column: int | None) -> int:
    value = row[column] if column is not None else None
    try:
        gender = int(value) if value is not None else 0
    except (TypeError, ValueError):
        return 0
    return gender if gender in (0, 1, 2) else 0


def _optional_text(row: tuple[Any, ...], column: int | None) -> str | None:
    if column is None or not row[column]:
        return None
    return str(row[column]).strip()
