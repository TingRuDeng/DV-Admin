# -*- coding: utf-8 -*-
"""用户导入导出双后端共享契约测试。"""

import base64
import csv
import io
from unittest.mock import patch

from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from drf_admin.apps.system.models import Departments, Permissions, Roles, Users
from drf_admin.apps.system.test_helpers import create_admin_user


def grant_permission(user: Users, code: str) -> None:
    permission, _ = Permissions.objects.get_or_create(
        perm=code,
        defaults={"name": code, "type": "BUTTON"},
    )
    user.roles.first().permissions.add(permission)
    cache.delete(f"user_info_{user.id}_perms")


def create_dept_scoped_user_context(permission_codes: tuple[str, ...] = ()):
    visible_dept = Departments.objects.create(name="导入范围内部门", status=1, sort=1)
    hidden_dept = Departments.objects.create(name="导入范围外部门", status=1, sort=2)
    role = Roles.objects.create(
        name="导入数据范围角色",
        code="import_scope_role",
        status=1,
        data_scope=Roles.DATA_SCOPE_DEPT,
    )
    for code in ("system:users:query", *permission_codes):
        permission, _ = Permissions.objects.get_or_create(
            perm=code,
            defaults={"name": code, "type": "BUTTON"},
        )
        role.permissions.add(permission)
    operator = Users.objects.create_user(
        username="import_scope_operator",
        password="admin123",
        dept=visible_dept,
        is_active=1,
    )
    operator.roles.add(role)
    visible_user = Users.objects.create_user(
        username="import_scope_visible",
        password="admin123",
        dept=visible_dept,
        is_active=1,
    )
    hidden_user = Users.objects.create_user(
        username="import_scope_hidden",
        password="admin123",
        dept=hidden_dept,
        is_active=1,
    )
    return {
        "operator": operator,
        "visible_dept": visible_dept,
        "hidden_dept": hidden_dept,
        "visible_user": visible_user,
        "hidden_user": hidden_user,
    }


def build_import_file(rows: list[list[str]]) -> SimpleUploadedFile:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        ["用户名*", "姓名", "邮箱", "手机号", "性别", "部门ID", "角色ID(多个用逗号分隔)"]
    )
    for row in rows:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return SimpleUploadedFile(
        "users.xlsx",
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class UserImportExportTestCase(TestCase):
    def setUp(self):
        cache.clear()
        self.client = APIClient()
        self.user = create_admin_user()
        self.client.force_authenticate(user=self.user)

    def test_template_returns_base64_xlsx_contract(self):
        grant_permission(self.user, "system:users:import")

        response = self.client.get("/api/v1/system/users/template")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        result = response.data["data"]
        self.assertEqual(result["filename"], "用户导入模板.xlsx")
        self.assertEqual(
            result["contentType"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(io.BytesIO(base64.b64decode(result["content"])), read_only=True)
        self.assertEqual(workbook.active["A1"].value, "用户名*")
        workbook.close()

    def test_export_follows_data_scope_and_field_masking(self):
        context = create_dept_scoped_user_context(("system:users:export",))
        visible_user = context["visible_user"]
        visible_user.email = "visible@example.com"
        visible_user.mobile = "13800138000"
        visible_user.save(update_fields=("email", "mobile"))
        context["hidden_user"].email = "hidden@example.com"
        context["hidden_user"].save(update_fields=("email",))
        self.client.force_authenticate(user=context["operator"])

        response = self.client.post("/api/v1/system/users/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        result = response.data["data"]
        self.assertEqual(result["contentType"], "text/csv;charset=utf-8")
        rows = list(
            csv.reader(
                io.StringIO(base64.b64decode(result["content"]).decode("utf-8-sig"))
            )
        )
        exported = {row[0]: row for row in rows[1:]}
        self.assertIn(visible_user.username, exported)
        self.assertNotIn(context["hidden_user"].username, exported)
        self.assertEqual(exported[visible_user.username][2], "v******@example.com")
        self.assertEqual(exported[visible_user.username][3], "138****8000")

    def test_import_creates_valid_user_with_roles(self):
        grant_permission(self.user, "system:users:import")
        grant_permission(self.user, "system:users:field:write")
        department = create_dept_scoped_user_context()["visible_dept"]
        role = Roles.objects.create(name="导入角色", code="import_role", status=1)
        uploaded = build_import_file(
            [
                [
                    "xlsx_import_user",
                    "导入用户",
                    "xlsx-import@example.com",
                    "13800138001",
                    "1",
                    str(department.id),
                    str(role.id),
                ]
            ]
        )

        response = self.client.post(
            f"/api/v1/system/users/import?deptId={department.id}",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["data"]["validCount"], 1)
        self.assertEqual(response.data["data"]["invalidCount"], 0)
        created = Users.objects.get(username="xlsx_import_user")
        self.assertEqual(created.dept_id, department.id)
        self.assertEqual(list(created.roles.values_list("id", flat=True)), [role.id])

    def test_import_rolls_back_when_second_save_fails(self):
        """意外保存失败时不能留下前面已写入的部分用户。"""
        grant_permission(self.user, "system:users:import")
        usernames = ["atomic_import_first", "atomic_import_second"]
        uploaded = build_import_file(
            [[username, "原子导入", "", "", "0", "", ""] for username in usernames]
        )
        original_create_user = Users.objects.create_user
        create_count = 0

        def fail_on_second_save(*args, **kwargs):
            nonlocal create_count
            create_count += 1
            if create_count == 2:
                raise RuntimeError("simulated persistence failure")
            return original_create_user(*args, **kwargs)

        with patch.object(Users.objects, "create_user", side_effect=fail_on_second_save):
            response = self.client.post(
                "/api/v1/system/users/import",
                {"file": uploaded},
                format="multipart",
            )

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertFalse(Users.objects.filter(username__in=usernames).exists())

    def test_import_rejects_department_outside_data_scope(self):
        context = create_dept_scoped_user_context(("system:users:import",))
        self.client.force_authenticate(user=context["operator"])
        uploaded = build_import_file(
            [
                [
                    "hidden_import_user",
                    "范围外导入",
                    "",
                    "",
                    "0",
                    str(context["hidden_dept"].id),
                    "",
                ]
            ]
        )

        response = self.client.post(
            "/api/v1/system/users/import",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["data"]["validCount"], 0)
        self.assertEqual(response.data["data"]["invalidCount"], 1)
        self.assertIn("目标部门超出", response.data["data"]["messageList"][0])
        self.assertFalse(Users.objects.filter(username="hidden_import_user").exists())

    def test_import_rejects_legacy_xls(self):
        grant_permission(self.user, "system:users:import")
        uploaded = SimpleUploadedFile("users.xls", b"legacy", content_type="application/vnd.ms-excel")

        response = self.client.post(
            "/api/v1/system/users/import",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("仅支持 .xlsx", str(response.data))

    @override_settings(MAX_UPLOAD_SIZE=4)
    def test_import_rejects_oversize_file_before_parsing(self):
        grant_permission(self.user, "system:users:import")
        uploaded = SimpleUploadedFile("users.xlsx", b"oversize")

        with patch(
            "drf_admin.apps.system.views.users.import_users",
        ) as import_mock:
            response = self.client.post(
                "/api/v1/system/users/import",
                {"file": uploaded},
                format="multipart",
            )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("文件大小不能超过 4 字节", str(response.data))
        import_mock.assert_not_called()
