# -*- coding: utf-8 -*-
"""
用户管理 Service
"""
from typing import List, Optional, Dict, Any, BinaryIO

from tortoise.expressions import Q

from app.core.cache import cache_service, CacheKeys
from app.core.config import settings
from app.core.exceptions import ValidationError, NotFound, BusinessError
from app.core.security import get_password_hash
from app.db.models.oauth import Users
from app.db.models.system import Departments, Roles
from app.schemas.base import PageResult
from app.schemas.system import (
    UserCreate,
    UserUpdate,
    UserOut,
    UserFormOut,
    UserPartialUpdate,
    UserImportResult,
)


class UserService:
    """用户管理服务"""

    async def _clear_user_cache(self, user_id: int) -> None:
        """
        清除用户缓存

        Args:
            user_id: 用户ID
        """
        # 清除用户权限缓存
        cache_key = CacheKeys.format_key(CacheKeys.USER_PERMISSIONS, user_id=user_id)
        await cache_service.delete(cache_key)
        # 清除用户菜单缓存
        cache_key = CacheKeys.format_key(CacheKeys.USER_MENUS, user_id=user_id)
        await cache_service.delete(cache_key)

    async def get_page(
        self,
        page: int,
        page_size: int,
        search: Optional[str] = None,
        is_active: Optional[int] = None,
        dept_id: Optional[int] = None,
    ) -> PageResult[UserOut]:
        """
        获取用户分页列表
        """
        # 构建查询条件
        query = Users.all()

        if search:
            query = query.filter(
                Q(username__icontains=search)
                | Q(name__icontains=search)
                | Q(mobile__icontains=search)
                | Q(email__icontains=search)
            )

        if is_active is not None:
            query = query.filter(is_active=is_active)

        if dept_id:
            query = query.filter(dept_id=dept_id)

        # 计算总数
        total = await query.count()

        # 分页查询 - 使用 prefetch_related 优化 N+1 查询
        users = await query.prefetch_related("roles", "dept").offset((page - 1) * page_size).limit(page_size).all()

        # 预加载所有部门（避免每个用户单独查询）
        dept_ids = [u.dept_id for u in users if u.dept_id]
        depts = {}
        if dept_ids:
            depts = {d.id: d.name for d in await Departments.filter(id__in=dept_ids).all()}

        # 序列化结果
        user_list = []
        for user in users:
            user_data = await self._serialize_user_optimized(user, depts)
            user_list.append(user_data)

        return PageResult.create(
            total=total,
            page=page,
            page_size=page_size,
            results=user_list,
        )

    async def get(self, user_id: int) -> UserOut:
        """
        获取用户详情
        """
        user = await Users.get_or_none(id=user_id)
        if not user:
            raise NotFound("用户不存在")

        return await self._serialize_user(user)

    async def get_form(self, user_id: int) -> UserFormOut:
        """
        获取用户表单详情（用于编辑回填）
        """
        user = await Users.get_or_none(id=user_id)
        if not user:
            raise NotFound("用户不存在")

        return await self._serialize_user_form(user)

    async def create(self, user_in: UserCreate) -> UserOut:
        """
        创建用户
        """
        # 检查用户名是否已存在
        existing = await Users.get_or_none(username=user_in.username)
        if existing:
            raise ValidationError("用户名已存在")

        # 检查手机号是否已存在
        if user_in.mobile:
            existing = await Users.get_or_none(mobile=user_in.mobile)
            if existing:
                raise ValidationError("手机号已存在")

        # 创建用户
        password = user_in.password or settings.default_password
        user = await Users.create(
            username=user_in.username,
            password=get_password_hash(password),
            name=user_in.name,
            email=user_in.email,
            mobile=user_in.mobile,
            avatar=user_in.avatar,
            gender=user_in.gender,
            is_active=user_in.is_active,
            dept_id=user_in.dept_id,
        )

        # 关联角色
        if user_in.role_ids:
            roles = await Roles.filter(id__in=user_in.role_ids).all()
            await user.roles.add(*roles)

        return await self._serialize_user(user)

    async def update(self, user_id: int, user_in: UserUpdate) -> UserOut:
        """
        更新用户
        """
        user = await Users.get_or_none(id=user_id)
        if not user:
            raise NotFound("用户不存在")

        # 检查手机号是否已被其他用户使用
        if user_in.mobile:
            existing = await Users.get_or_none(mobile=user_in.mobile)
            if existing and existing.id != user_id:
                raise ValidationError("手机号已存在")

        # 更新用户字段
        update_fields = {}
        if user_in.name is not None:
            update_fields["name"] = user_in.name
        if user_in.email is not None:
            update_fields["email"] = user_in.email
        if user_in.mobile is not None:
            update_fields["mobile"] = user_in.mobile
        if user_in.gender is not None:
            update_fields["gender"] = user_in.gender
        if user_in.is_active is not None:
            update_fields["is_active"] = user_in.is_active
        if user_in.dept_id is not None:
            update_fields["dept_id"] = user_in.dept_id
        if user_in.avatar is not None:
            update_fields["avatar"] = user_in.avatar

        if update_fields:
            await Users.filter(id=user_id).update(**update_fields)
            await user.refresh_from_db()

        # 更新角色关联
        if user_in.role_ids is not None:
            # 清除现有角色
            await user.roles.clear()
            # 添加新角色
            if user_in.role_ids:
                roles = await Roles.filter(id__in=user_in.role_ids).all()
                await user.roles.add(*roles)

        # 清除用户缓存（角色变更会影响权限和菜单）
        await self._clear_user_cache(user_id)

        return await self._serialize_user(user)

    async def partial_update(self, user_id: int, user_in: UserPartialUpdate) -> UserOut:
        """
        局部更新用户（状态）
        """
        user = await Users.get_or_none(id=user_id)
        if not user:
            raise NotFound("用户不存在")

        # 更新状态
        user.is_active = user_in.is_active
        await user.save()

        # 清除用户缓存
        await self._clear_user_cache(user_id)

        return await self._serialize_user(user)

    async def delete(self, user_id: int, current_user_id: int) -> None:
        """
        删除用户
        """
        user = await Users.get_or_none(id=user_id)
        if not user:
            raise NotFound("用户不存在")

        # 不能删除自己
        if user.id == current_user_id:
            raise BusinessError("不能删除当前登录用户")

        # 删除用户（级联删除角色关联）
        await user.delete()

        # 清除用户缓存
        await self._clear_user_cache(user_id)

    async def batch_delete(self, ids: List[int], current_user_id: int) -> None:
        """
        批量删除用户
        """
        if current_user_id in ids:
            raise BusinessError("不能删除当前登录用户")

        # 批量删除
        await Users.filter(id__in=ids).delete()

        # 清除所有被删除用户的缓存
        for user_id in ids:
            await self._clear_user_cache(user_id)

    async def get_options(self) -> List[Dict[str, Any]]:
        """
        获取用户下拉选项
        """
        users = await Users.filter(is_active=1).all()

        return [
            {
                "value": user.id,
                "label": f"{user.name or user.username}({user.username})",
            }
            for user in users
        ]

    async def reset_password(self, user_id: int) -> None:
        """
        重置用户密码
        """
        user = await Users.get_or_none(id=user_id)
        if not user:
            raise NotFound("用户不存在")

        default_password = settings.default_password
        user.password = get_password_hash(default_password)
        await user.save()

    async def import_users(
        self, file: BinaryIO, dept_id: Optional[int] = None
    ) -> UserImportResult:
        """
        导入用户

        Args:
            file: Excel 文件对象
            dept_id: 默认部门ID

        Returns:
            导入结果
        """
        from openpyxl import load_workbook

        # 读取 Excel 文件
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            raise ValidationError(f"Excel 文件解析失败: {str(e)}")

        # 获取表头
        headers = [cell.value for cell in ws[1]]
        if not headers or headers[0] is None:
            raise ValidationError("Excel 文件格式错误，缺少表头")

        # 验证必要字段
        required_fields = ["用户名*"]
        for field in required_fields:
            if field not in headers:
                raise ValidationError(f"Excel 文件缺少必要字段: {field}")

        # 获取列索引
        username_idx = headers.index("用户名*")
        name_idx = headers.index("姓名") if "姓名" in headers else None
        email_idx = headers.index("邮箱") if "邮箱" in headers else None
        mobile_idx = headers.index("手机号") if "手机号" in headers else None
        gender_idx = headers.index("性别") if "性别" in headers else None
        dept_idx = headers.index("部门ID") if "部门ID" in headers else None
        role_idx = headers.index("角色ID(多个用逗号分隔)") if "角色ID(多个用逗号分隔)" in headers else None

        # 统计结果
        valid_count = 0
        invalid_count = 0
        message_list = []

        # 预加载所有部门和角色
        all_depts = {dept.id: dept for dept in await Departments.all()}
        all_roles = {role.id: role for role in await Roles.filter(status=1).all()}

        # 预加载已存在的用户名和手机号
        existing_usernames = set(
            await Users.all().values_list("username", flat=True)
        )
        existing_mobiles = set(
            await Users.filter(mobile__isnull=False).values_list("mobile", flat=True)
        )

        # 批量创建用户列表
        users_to_create = []

        # 从第二行开始读取数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 获取用户名（必填）
                username = str(row[username_idx]).strip() if row[username_idx] else None
                if not username:
                    message_list.append(f"第{row_idx}行: 用户名不能为空")
                    invalid_count += 1
                    continue

                # 检查用户名是否已存在
                if username in existing_usernames:
                    message_list.append(f"第{row_idx}行: 用户名 '{username}' 已存在")
                    invalid_count += 1
                    continue

                # 获取其他字段
                name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else None
                email = str(row[email_idx]).strip() if email_idx is not None and row[email_idx] else None
                mobile = str(row[mobile_idx]).strip() if mobile_idx is not None and row[mobile_idx] else None

                # 检查手机号是否已存在
                if mobile and mobile in existing_mobiles:
                    message_list.append(f"第{row_idx}行: 手机号 '{mobile}' 已存在")
                    invalid_count += 1
                    continue

                # 处理性别
                gender = 0
                if gender_idx is not None and row[gender_idx] is not None:
                    try:
                        gender = int(row[gender_idx])
                        if gender not in [0, 1, 2]:
                            gender = 0
                    except (ValueError, TypeError):
                        gender = 0

                # 处理部门ID
                user_dept_id = dept_id  # 默认使用传入的部门ID
                if dept_idx is not None and row[dept_idx] is not None:
                    try:
                        import_dept_id = int(row[dept_idx])
                        if import_dept_id in all_depts:
                            user_dept_id = import_dept_id
                        else:
                            message_list.append(f"第{row_idx}行: 部门ID '{import_dept_id}' 不存在，使用默认部门")
                    except (ValueError, TypeError):
                        message_list.append(f"第{row_idx}行: 部门ID格式错误，使用默认部门")

                # 处理角色ID
                role_ids = []
                if role_idx is not None and row[role_idx] is not None:
                    try:
                        role_str = str(row[role_idx]).strip()
                        if role_str:
                            role_ids = [
                                int(rid.strip())
                                for rid in role_str.split(",")
                                if rid.strip()
                            ]
                            # 验证角色是否存在
                            invalid_roles = [
                                rid for rid in role_ids if rid not in all_roles
                            ]
                            if invalid_roles:
                                message_list.append(
                                    f"第{row_idx}行: 角色ID {invalid_roles} 不存在或已禁用"
                                )
                                role_ids = [rid for rid in role_ids if rid in all_roles]
                    except (ValueError, TypeError):
                        message_list.append(f"第{row_idx}行: 角色ID格式错误")

                # 创建用户对象
                user = Users(
                    username=username,
                    password=get_password_hash(settings.default_password),
                    name=name or username,
                    email=email,
                    mobile=mobile,
                    gender=gender,
                    is_active=1,
                    dept_id=user_dept_id,
                    avatar="avatar/default.png",
                )
                users_to_create.append((user, role_ids))

                # 添加到已存在集合，防止重复
                existing_usernames.add(username)
                if mobile:
                    existing_mobiles.add(mobile)

                valid_count += 1

            except Exception as e:
                message_list.append(f"第{row_idx}行: 数据处理错误 - {str(e)}")
                invalid_count += 1

        # 批量创建用户
        for user, role_ids in users_to_create:
            await user.save()
            # 关联角色
            if role_ids:
                roles = [all_roles[rid] for rid in role_ids if rid in all_roles]
                await user.roles.add(*roles)

        return UserImportResult(
            valid_count=valid_count,
            invalid_count=invalid_count,
            message_list=message_list,
        )

    async def get_import_template(self) -> Dict[str, Any]:
        """
        获取用户导入模板（Excel格式）
        """
        import io
        import base64
        from openpyxl import Workbook

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "用户导入模板"

        # 设置表头
        headers = ["用户名*", "姓名", "邮箱", "手机号", "性别", "部门ID", "角色ID(多个用逗号分隔)"]
        ws.append(headers)

        # 设置示例数据
        example_data = ["zhangsan", "张三", "zhangsan@example.com", "13800138000", "1", "1", "1,2"]
        ws.append(example_data)

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 25

        # 保存到字节流
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # 转换为 base64
        b64_content = base64.b64encode(buffer.getvalue()).decode("utf-8")

        return {
            "filename": "用户导入模板.xlsx",
            "content": b64_content,
        }

    async def export_users(self) -> Dict[str, Any]:
        """
        导出用户
        """
        import io
        import csv
        import base64

        users = await Users.all()

        headers = ["用户名", "姓名", "邮箱", "手机号", "状态", "创建时间"]
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)

        for user in users:
            writer.writerow([
                user.username,
                user.name or "",
                user.email or "",
                user.mobile or "",
                "启用" if user.is_active else "禁用",
                user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
            ])

        content = buffer.getvalue()
        b64_content = base64.b64encode(content.encode("utf-8")).decode("utf-8")

        return {
            "filename": "用户导出.csv",
            "content": b64_content,
        }

    async def _serialize_user(self, user: Users) -> UserOut:
        """
        序列化用户对象（处理关联字段）
        """
        # 获取部门名称
        dept_name = None
        if user.dept_id:
            department = await Departments.get_or_none(id=user.dept_id)
            if department:
                dept_name = department.name

        # 获取角色名称
        await user.fetch_related("roles")
        role_names = ",".join([role.name for role in user.roles])

        return UserOut(
            id=user.id,
            username=user.username,
            name=user.name,
            email=user.email,
            mobile=user.mobile,
            avatar=user.avatar,
            gender=user.gender,
            is_active=user.is_active,
            dept_id=user.dept_id,
            dept_name=dept_name,
            role_names=role_names,
            created_at=user.created_at,
            updated_at=user.updated_at,
        )

    async def _serialize_user_optimized(self, user: Users, depts: Dict[int, str]) -> UserOut:
        """
        序列化用户对象（优化版本，避免 N+1 查询）

        Args:
            user: 用户对象
            depts: 预加载的部门字典 {dept_id: dept_name}
        """
        # 从预加载的部门字典获取部门名称
        dept_name = depts.get(user.dept_id) if user.dept_id else None

        # 角色已通过 prefetch_related 加载，直接访问
        role_names = ",".join([role.name for role in user.roles])

        return UserOut(
            id=user.id,
            username=user.username,
            name=user.name,
            email=user.email,
            mobile=user.mobile,
            avatar=user.avatar,
            gender=user.gender,
            is_active=user.is_active,
            dept_id=user.dept_id,
            dept_name=dept_name,
            role_names=role_names,
            created_at=user.created_at,
            updated_at=user.updated_at,
        )

    async def _serialize_user_form(self, user: Users) -> UserFormOut:
        """
        序列化用户对象（表单回填用）
        """
        base = await self._serialize_user(user)

        await user.fetch_related("roles")
        role_ids = [role.id for role in user.roles]

        return UserFormOut(**base.model_dump(), roles=role_ids)


# 导出服务实例
user_service = UserService()
